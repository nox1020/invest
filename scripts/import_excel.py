"""Reset database and import portfolio data from invest.xlsx carefully."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from app.bootstrap import bootstrap
from app.config import (
    CALENDAR_JALALI,
    CURRENCY_TOMAN,
    TRADE_STATUS_CLOSED,
    TRADE_STATUS_OPEN,
    THEME_DARK,
)
from app.models.asset import Asset
from app.models.trade import Trade
from app.utils import calc
from app.utils.dates import now_iso, today_iso

EXCEL_PATH = Path(r"c:\Users\Aseman\OneDrive\Desktop\1405.03.19\invest.xlsx")
USDT_RATE = 189_000.0  # تومان به ازای هر USDT (از ردیف قیمت USDT در فایل)
SHIB_PRICE_USDT = 4.28e-06
GOLD_PRICE = 18_000_000.0


def _iso(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)[:10]


def _clear_db(conn) -> None:
    conn.execute("DELETE FROM trades")
    conn.execute("DELETE FROM capital_snapshots")
    conn.execute("DELETE FROM assets")
    conn.commit()


def import_excel() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["داشبورد"]

    # --- Prices from header ---
    usdt_rate = float(ws.cell(1, 2).value or USDT_RATE)
    shib_px_usdt = float(ws.cell(2, 2).value or SHIB_PRICE_USDT)
    gold_px = float(ws.cell(3, 2).value or GOLD_PRICE)
    shib_px_toman = shib_px_usdt * usdt_rate

    print(f"USDT rate: {usdt_rate:,.0f} Toman")
    print(f"SHIB price: {shib_px_usdt} USDT = {shib_px_toman:.6f} Toman")
    print(f"GOLD price: {gold_px:,.0f} Toman")

    ctx = bootstrap()
    _clear_db(ctx.conn)
    print("Database cleared.")

    # Create assets
    assets: dict[str, Asset] = {}
    for name, symbol, price in [
        ("تتر", "USDT", usdt_rate),
        ("شیبا", "SHIB", shib_px_toman),
        ("طلا", "GOLD", gold_px),
        ("عیار - 200", "AYAR200", 1.0),
    ]:
        asset = ctx.trades.create_asset(
            name=name,
            symbol=symbol,
            quantity=0.0,
            avg_buy_price=0.0,
            current_price=price,
            notes="وارد شده از invest.xlsx",
        )
        assets[symbol] = asset

    open_rows = []
    for r in range(5, 18):
        buy_date = ws.cell(r, 1).value
        buy_tot = ws.cell(r, 4).value
        cur_tot = ws.cell(r, 5).value
        name = ws.cell(r, 6).value
        qty = ws.cell(r, 7).value
        note = ws.cell(r, 10).value or ""
        if not buy_date or not qty or buy_tot is None:
            continue
        # ردیف ۵ جمع کل شیباهای باز است (مجموع ردیف‌های ۷ تا ۱۰)، معامله جدا نیست
        if r == 5 and name is None:
            print(f"Skipping aggregate SHIB summary row {r}")
            continue
        raw_name = (str(name).strip().upper() if name else "SHIB")
        if raw_name in ("SHIB", "SHIB?"):
            symbol = "SHIB"
            # مبلغ کل در فایل به USDT است
            buy_price = (float(buy_tot) / float(qty)) * usdt_rate
            current_price = shib_px_toman
        elif raw_name == "GOLD":
            symbol = "GOLD"
            buy_price = float(buy_tot) / float(qty)
            current_price = gold_px
        else:
            raise ValueError(f"Unknown open asset at row {r}: {name!r}")
        open_rows.append(
            {
                "row": r,
                "symbol": symbol,
                "qty": float(qty),
                "buy_price": buy_price,
                "buy_date": _iso(buy_date),
                "note": str(note).strip(),
                "current_price": current_price,
                "buy_tot_raw": float(buy_tot),
                "cur_tot_raw": float(cur_tot) if cur_tot is not None else None,
            }
        )

    print(f"Open trades to import: {len(open_rows)}")
    for item in open_rows:
        ctx.trades.register_buy(
            asset_id=assets[item["symbol"]].id,
            quantity=item["qty"],
            buy_price=item["buy_price"],
            buy_fee=0.0,
            buy_date=item["buy_date"],
            buy_note=item["note"] or f"ورود از اکسل ردیف {item['row']}",
            current_price=item["current_price"],
        )
        print(
            f"  OPEN {item['symbol']} qty={item['qty']} "
            f"buy={item['buy_price']:.8f} date={item['buy_date']}"
        )

    # Closed trades
    closed_specs = []
    for r in range(22, 26):
        buy_date = ws.cell(r, 1).value
        sell_date = ws.cell(r, 2).value
        buy_tot = ws.cell(r, 4).value
        sell_tot = ws.cell(r, 5).value
        name = ws.cell(r, 6).value
        typ = ws.cell(r, 7).value
        pnl_sheet = ws.cell(r, 9).value
        note = ws.cell(r, 10).value or ""
        if not buy_date or buy_tot is None or sell_tot is None or not name:
            continue
        raw = str(name).strip().upper()
        if raw == "GOLD":
            symbol = "GOLD"
            # typ like '47.930 G ' or 2.48
            if isinstance(typ, str):
                qty = float(str(typ).upper().replace("G", "").strip())
            else:
                qty = float(typ)
            buy_price = float(buy_tot) / qty
            sell_price = float(sell_tot) / qty
        elif raw == "SHIB":
            symbol = "SHIB"
            qty = float(typ)
            buy_price = (float(buy_tot) / qty) * usdt_rate
            sell_price = (float(sell_tot) / qty) * usdt_rate
        elif "عیار" in str(name):
            symbol = "AYAR200"
            # یک موقعیت کامل با مبالغ تومانی
            qty = 1.0
            buy_price = float(buy_tot)
            sell_price = float(sell_tot)
        else:
            raise ValueError(f"Unknown closed asset at row {r}: {name!r}")

        closed_specs.append(
            {
                "row": r,
                "symbol": symbol,
                "qty": qty,
                "buy_price": buy_price,
                "sell_price": sell_price,
                "buy_date": _iso(buy_date),
                "sell_date": _iso(sell_date),
                "note": str(note).strip(),
                "pnl_sheet": float(pnl_sheet) if pnl_sheet is not None else None,
            }
        )

    print(f"Closed trades to import: {len(closed_specs)}")
    for item in closed_specs:
        asset = assets[item["symbol"]]
        # Insert closed trade directly (inventory already reflects only open lots)
        pnl = calc.realized_pnl(
            item["qty"], item["buy_price"], item["sell_price"], 0.0, 0.0
        )
        pct = calc.return_pct(
            item["qty"], item["buy_price"], item["sell_price"], 0.0, 0.0
        )
        days = calc.holding_days(item["buy_date"], item["sell_date"])
        trade = Trade(
            id=None,
            asset_id=asset.id,  # type: ignore[arg-type]
            status=TRADE_STATUS_CLOSED,
            quantity=item["qty"],
            buy_price=item["buy_price"],
            buy_fee=0.0,
            buy_date=item["buy_date"],
            buy_note=item["note"] or f"بسته از اکسل ردیف {item['row']}",
            sell_price=item["sell_price"],
            sell_fee=0.0,
            sell_date=item["sell_date"],
            sell_note=item["note"],
            realized_pnl=pnl,
            return_pct=pct,
            holding_days=days,
        )
        ctx.trades.trades.create(trade)
        print(
            f"  CLOSED {item['symbol']} qty={item['qty']} "
            f"pnl={pnl:,.0f} (sheet={item['pnl_sheet']}) "
            f"{item['buy_date']} -> {item['sell_date']}"
        )

    # Refresh asset quantities/prices from open lots only
    for symbol, asset in assets.items():
        fresh = ctx.portfolio.assets.get(asset.id)  # type: ignore[arg-type]
        if not fresh:
            continue
        if symbol == "USDT":
            # در فایل موجودی نقد تتر مشخص نیست؛ فقط قیمت مرجع نگه داشته می‌شود
            fresh.quantity = 0.0
            fresh.avg_buy_price = 0.0
            fresh.current_price = usdt_rate
        elif symbol == "AYAR200":
            fresh.quantity = 0.0
            fresh.avg_buy_price = 0.0
            fresh.current_price = 1.0
        elif symbol == "SHIB":
            fresh.current_price = shib_px_toman
        elif symbol == "GOLD":
            fresh.current_price = gold_px
        ctx.portfolio.assets.update(fresh)

    # Settings
    ctx.settings.calendar = CALENDAR_JALALI
    ctx.settings.currency = CURRENCY_TOMAN
    ctx.settings.theme = THEME_DARK
    ctx.settings.first_run_done = True
    ctx.save_settings()

    # Snapshots for chart
    value = ctx.portfolio.record_snapshot()
    metrics = ctx.portfolio.get_metrics()

    print("--- SUMMARY ---")
    print(f"Assets: {len(ctx.portfolio.assets.list_all())}")
    print(f"Open trades: {metrics.open_count}")
    print(f"Closed trades: {metrics.closed_count}")
    print(f"Portfolio value: {metrics.total_value:,.0f}")
    print(f"Unrealized PnL: {metrics.unrealized_pnl:,.0f}")
    print(f"Realized PnL: {metrics.realized_pnl:,.0f}")
    for a in ctx.portfolio.assets.list_all():
        print(
            f"  {a.name} ({a.symbol}): qty={a.quantity} "
            f"avg={a.avg_buy_price:.8f} px={a.current_price:.8f} "
            f"value={a.total_value:,.0f}"
        )

    # Verify SHIB open qty vs excel
    shib_qty_excel = sum(i["qty"] for i in open_rows if i["symbol"] == "SHIB")
    gold_qty_excel = sum(i["qty"] for i in open_rows if i["symbol"] == "GOLD")
    shib = next(a for a in ctx.portfolio.assets.list_all() if a.symbol == "SHIB")
    gold = next(a for a in ctx.portfolio.assets.list_all() if a.symbol == "GOLD")
    assert abs(shib.quantity - shib_qty_excel) < 1e-6, (shib.quantity, shib_qty_excel)
    assert abs(gold.quantity - gold_qty_excel) < 1e-6, (gold.quantity, gold_qty_excel)
    assert metrics.open_count == 11, metrics.open_count
    assert metrics.closed_count == 4, metrics.closed_count
    print("Verification OK.")
    print(f"Snapshot value today: {value:,.2f}")
    ctx.close()


if __name__ == "__main__":
    import_excel()
