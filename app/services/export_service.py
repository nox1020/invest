"""Lazy export helpers so core bootstrap does not require Excel/PDF deps at import time."""

from __future__ import annotations

from pathlib import Path

from app.models.asset import Asset
from app.models.trade import Trade


class ExportService:
    def export_excel(
        self,
        path: Path,
        *,
        assets: list[Asset],
        open_trades: list[Trade],
        closed_trades: list[Trade],
        calendar: str,
        currency: str,
        fx_rate: float | None = None,
    ) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        from app.utils import calc
        from app.utils.dates import format_short_date
        from app.utils.money import to_display_amount

        def amt(value: float) -> float:
            return to_display_amount(value, currency, fx_rate=fx_rate)

        wb = Workbook()
        ws_assets = wb.active
        ws_assets.title = "Assets"
        ws_assets.append(
            ["Name", "Symbol", "Quantity", "Avg Buy", "Current", "Value", "PnL", "PnL %"]
        )
        for a in assets:
            ws_assets.append(
                [
                    a.name,
                    a.symbol,
                    a.quantity,
                    amt(a.avg_buy_price),
                    amt(a.current_price),
                    amt(a.total_value),
                    amt(a.unrealized_pnl),
                    a.unrealized_pnl_pct,
                ]
            )

        ws_open = wb.create_sheet("Open Trades")
        ws_open.append(
            [
                "Asset",
                "Buy Date",
                "Qty",
                "Buy Price",
                "Fee",
                "Current",
                "Value",
                "PnL",
                "Note",
            ]
        )
        for t in open_trades:
            pnl = calc.unrealized_pnl(
                t.quantity, t.buy_price, t.current_price, t.buy_fee
            )
            ws_open.append(
                [
                    t.asset_name,
                    format_short_date(t.buy_date, calendar),
                    t.quantity,
                    amt(t.buy_price),
                    amt(t.buy_fee),
                    amt(t.current_price),
                    amt(t.quantity * t.current_price),
                    amt(pnl),
                    t.buy_note,
                ]
            )

        ws_closed = wb.create_sheet("Closed Trades")
        ws_closed.append(
            [
                "Asset",
                "Buy Date",
                "Sell Date",
                "Qty",
                "Buy",
                "Sell",
                "PnL",
                "Return %",
                "Days",
            ]
        )
        for t in closed_trades:
            ws_closed.append(
                [
                    t.asset_name,
                    format_short_date(t.buy_date, calendar),
                    format_short_date(t.sell_date, calendar),
                    t.quantity,
                    amt(t.buy_price),
                    amt(t.sell_price or 0),
                    amt(t.realized_pnl or 0),
                    t.return_pct,
                    t.holding_days,
                ]
            )

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)

        path = Path(path)
        wb.save(path)
        return path

    def export_pdf(
        self,
        path: Path,
        *,
        assets: list[Asset],
        open_trades: list[Trade],
        closed_trades: list[Trade],
        calendar: str,
        currency: str,
        title: str | None = None,
        fx_rate: float | None = None,
    ) -> Path:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        from app.config import APP_NAME
        from app.utils.dates import format_short_date
        from app.utils.money import format_money, format_pct

        path = Path(path)
        doc = SimpleDocTemplate(str(path), pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story = [Paragraph(title or APP_NAME, styles["Title"]), Spacer(1, 12)]

        def money(value: float, *, show_sign: bool = False) -> str:
            return format_money(
                value, currency, show_sign=show_sign, fx_rate=fx_rate
            )

        def _table(headers, rows):
            data = [headers] + rows
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E40")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F3F5F4")],
                        ),
                    ]
                )
            )
            return table

        story.append(Paragraph("Assets", styles["Heading2"]))
        asset_rows = [
            [
                a.name,
                a.symbol,
                f"{a.quantity}",
                money(a.avg_buy_price),
                money(a.current_price),
                money(a.total_value),
                money(a.unrealized_pnl, show_sign=True),
            ]
            for a in assets
        ]
        story.append(
            _table(
                ["Name", "Symbol", "Qty", "Avg Buy", "Current", "Value", "PnL"],
                asset_rows or [["—", "", "", "", "", "", ""]],
            )
        )
        story.append(Spacer(1, 16))

        story.append(Paragraph("Open Trades", styles["Heading2"]))
        open_rows = [
            [
                t.asset_name,
                format_short_date(t.buy_date, calendar),
                f"{t.quantity}",
                money(t.buy_price),
                money(t.current_price),
            ]
            for t in open_trades
        ]
        story.append(
            _table(
                ["Asset", "Buy Date", "Qty", "Buy", "Current"],
                open_rows or [["—", "", "", "", ""]],
            )
        )
        story.append(Spacer(1, 16))

        story.append(Paragraph("Closed Trades", styles["Heading2"]))
        closed_rows = [
            [
                t.asset_name,
                format_short_date(t.buy_date, calendar),
                format_short_date(t.sell_date, calendar),
                f"{t.quantity}",
                money(t.realized_pnl or 0, show_sign=True),
                format_pct(t.return_pct or 0),
            ]
            for t in closed_trades
        ]
        story.append(
            _table(
                ["Asset", "Buy", "Sell", "Qty", "PnL", "Return"],
                closed_rows or [["—", "", "", "", "", ""]],
            )
        )

        doc.build(story)
        return path
