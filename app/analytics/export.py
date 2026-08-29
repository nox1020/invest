"""Export analytics reports to CSV / Excel (shared structure)."""

from __future__ import annotations

import csv
from pathlib import Path

from app.analytics.models import AnalyticsBundle, PortfolioSummaryReport
from app.analytics.report_generator import AnalyticsReportGenerator


def export_summary_csv(path: Path, report: PortfolioSummaryReport) -> Path:
    cap = report.capital
    tr = report.trades
    rows = [
        ("metric", "value"),
        ("portfolio_value", cap.portfolio_value),
        ("cash_invested", cap.cash_invested),
        ("realized_pnl", cap.realized_pnl),
        ("unrealized_pnl", cap.unrealized_pnl),
        ("total_pnl", cap.total_pnl),
        ("roi_pct", cap.roi_pct),
        ("total_fees", cap.total_fees),
        ("win_rate_pct", tr.win_rate_pct),
        ("profit_factor", tr.profit_factor),
        ("expectancy", tr.expectancy),
        ("generated_at", report.generated_at),
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return path


def export_bundle_csv(path: Path, bundle: AnalyticsBundle) -> Path:
    """Multi-section CSV: summary + assets + monthly periods."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["section", "portfolio_summary"])
        cap = bundle.capital
        w.writerow(["portfolio_value", cap.portfolio_value])
        w.writerow(["total_pnl", cap.total_pnl])
        w.writerow(["roi_pct", cap.roi_pct])
        w.writerow([])
        w.writerow(["section", "assets"])
        w.writerow(
            [
                "name",
                "allocation_pct",
                "current_value",
                "return_pct",
                "realized",
                "unrealized",
            ]
        )
        for a in bundle.assets:
            w.writerow(
                [
                    a.name,
                    a.allocation_pct,
                    a.current_value,
                    a.return_pct,
                    a.realized_pnl,
                    a.unrealized_pnl,
                ]
            )
        w.writerow([])
        w.writerow(["section", "monthly"])
        w.writerow(["period", "net_pnl", "trades", "success_rate_pct"])
        for p in bundle.periods.get("monthly", []):
            w.writerow([p.key, p.net_pnl, p.trade_count, p.success_rate_pct])
    return path


def export_analytics_excel(
    path: Path,
    generator: AnalyticsReportGenerator,
    *,
    calendar: str,
) -> Path:
    from openpyxl import Workbook

    summary = generator.portfolio_summary(calendar=calendar)
    perf = generator.performance_report(calendar=calendar)
    assets = generator.asset_report(calendar=calendar)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    cap = summary.capital
    for label, val in (
        ("Portfolio Value", cap.portfolio_value),
        ("Cash Invested", cap.cash_invested),
        ("Total PnL", cap.total_pnl),
        ("ROI %", cap.roi_pct),
        ("Win Rate %", summary.trades.win_rate_pct),
        ("Profit Factor", summary.trades.profit_factor),
    ):
        ws.append([label, val])

    ws_a = wb.create_sheet("Assets")
    ws_a.append(
        ["Name", "Allocation %", "Value", "Return %", "Realized", "Unrealized"]
    )
    for a in assets.rows:
        ws_a.append(
            [
                a.name,
                a.allocation_pct,
                a.current_value,
                a.return_pct,
                a.realized_pnl,
                a.unrealized_pnl,
            ]
        )

    ws_m = wb.create_sheet("Monthly")
    ws_m.append(["Period", "Net PnL", "Trades", "Success %"])
    for p in perf.periods.get("monthly", []):
        ws_m.append([p.key, p.net_pnl, p.trade_count, p.success_rate_pct])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
